import pandas as pd
import re
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO


def ler_output(caminho_arquivo: str) -> pd.DataFrame:
    with open(caminho_arquivo, "r", encoding="utf-8") as f:
        conteudo = f.read()

    padrao = r"===\s*(.+?)\s*===\s*Size,Operations\n([\s\S]*?)(?=\n===|$)"
    blocos = re.findall(padrao, conteudo)

    registros = []
    for nome_alg, dados in blocos:
        linhas = [l.strip() for l in dados.strip().splitlines() if l.strip()]
        for linha in linhas:
            if "," not in linha:
                continue
            size, ops = linha.split(",", 1)
            registros.append({
                "Algoritmo": nome_alg.strip(),
                "Tamanho": int(size),
                "Operações": int(ops)
            })

    df = pd.DataFrame(registros)
    return df


def gerar_grafico(df: pd.DataFrame) -> BytesIO:
    plt.figure(figsize=(10, 6))
    for alg in df["Algoritmo"].unique():
        subset = df[df["Algoritmo"] == alg]
        plt.plot(subset["Tamanho"], subset["Operações"], label=alg)

    plt.xlabel("Tamanho do vetor")
    plt.ylabel("Operações médias")
    plt.title("Comparação da complexidade dos algoritmos de ordenação")
    plt.legend()
    plt.grid(True)

    buffer = BytesIO()
    plt.savefig(buffer, format="png", dpi=150, bbox_inches="tight")
    buffer.seek(0)
    plt.close()
    return buffer


def gerar_grafico_individual(df: pd.DataFrame, algoritmo: str) -> BytesIO:
    subset = df[df["Algoritmo"] == algoritmo]
    plt.figure(figsize=(8, 5))
    plt.plot(subset["Tamanho"], subset["Operações"], marker="o")
    plt.xlabel("Tamanho do vetor")
    plt.ylabel("Operações médias")
    plt.title(f"Complexidade de {algoritmo}")
    plt.grid(True)

    buffer = BytesIO()
    plt.savefig(buffer, format="png", dpi=150, bbox_inches="tight")
    buffer.seek(0)
    plt.close()
    return buffer


def gerar_planilha(df: pd.DataFrame, caminho_saida: str = "resultados.xlsx"):
    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Geral")

        for alg in df["Algoritmo"].unique():
            df_alg = df[df["Algoritmo"] == alg]
            df_alg.to_excel(writer, index=False, sheet_name=alg[:30])

    wb = load_workbook(caminho_saida)

    ws_geral = wb.create_sheet("Gráfico Geral")
    img_buffer = gerar_grafico(df)
    img = XLImage(img_buffer)
    img.anchor = "A1"
    ws_geral.add_image(img)

    for alg in df["Algoritmo"].unique():
        ws = wb[alg[:30]]
        img_buffer = gerar_grafico_individual(df, alg)
        img = XLImage(img_buffer)
        img.anchor = "E2" 
        ws.add_image(img)

    wb.save(caminho_saida)
    print(f" planilha '{caminho_saida}' gerada")


def main():
    arquivo_entrada = "output.txt"
    arquivo_saida = "resultados.xlsx"

    df = ler_output(arquivo_entrada)
    gerar_planilha(df, arquivo_saida)


if __name__ == "__main__":
    main()
